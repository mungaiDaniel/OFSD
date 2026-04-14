"""
Reports & Statements Module

Endpoints for generating investor statements, portfolio reports based on ValuationRun and EpochLedger.
All financial values formatted with 2-decimal rounding.
Only 'Committed' valuation runs appear in reports.
"""

from flask import request, Blueprint, jsonify, send_file
from flask_jwt_extended import jwt_required
from sqlalchemy import func, case, and_, or_, desc, asc
from decimal import Decimal
from io import BytesIO
import pandas as pd
from datetime import datetime, timezone
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT

from app.database.database import db
from app.Investments.model import Investment, EpochLedger
from app.Batch.model import Batch
from app.Batch.core_fund import CoreFund
from app.Valuation.model import ValuationRun, Statement
from typing import Optional
import traceback


reports_v1 = Blueprint("reports_v1", __name__, url_prefix="/")


def float_2dp(val) -> float:
    """Round to 2 decimal places using Decimal for precision."""
    if val is None:
        return 0.0
    return float(Decimal(str(val)).quantize(Decimal("0.01")))


def format_currency(val) -> str:
    """Format value as currency with 2 decimal places and thousands separator."""
    return f"{float_2dp(val):,.2f}"




def verify_ledger_hash_chain(investor_code: str, fund_name: str) -> dict:
    """
    Verify the cryptographic integrity of an investor's ledger record chain.
    Returns verification status and any broken links.
    """
    try:
        # Get all ledger entries for this investor/fund in chronological order
        entries = db.session.query(EpochLedger).filter(
            EpochLedger.internal_client_code == investor_code,
            func.lower(EpochLedger.fund_name) == fund_name.lower()
        ).order_by(EpochLedger.epoch_end.asc()).all()

        if not entries:
            return {"valid": True, "message": "No ledger entries found", "entries_checked": 0}

        broken_links = []
        expected_previous_hash = GENESIS_HASH

        for entry in entries:
            # Recalculate what the hash should be
            payload = _ledger_hash_payload(
                internal_client_code=entry.internal_client_code,
                fund_name=entry.fund_name,
                epoch_start=entry.epoch_start,
                epoch_end=entry.epoch_end,
                performance_rate=entry.performance_rate,
                start_balance=entry.start_balance,
                deposits=entry.deposits,
                withdrawals=entry.withdrawals,
                profit=entry.profit,
                end_balance=entry.end_balance,
                previous_hash=expected_previous_hash,
            )
            recalculated_hash = _sha256_hex(payload)

            # Check if stored hash matches recalculated hash
            if entry.current_hash != recalculated_hash:
                broken_links.append({
                    "epoch_end": entry.epoch_end.isoformat(),
                    "stored_hash": entry.current_hash,
                    "expected_hash": recalculated_hash
                })

            # Update expected previous hash for next iteration
            expected_previous_hash = entry.current_hash

        return {
            "valid": len(broken_links) == 0,
            "message": "Hash chain verification complete",
            "entries_checked": len(entries),
            "broken_links": broken_links
        }

    except Exception as e:
        return {
            "valid": False,
            "message": f"Error during verification: {str(e)}",
            "entries_checked": 0,
            "broken_links": []
        }


def _ledger_hash_payload(
    *,
    internal_client_code: str,
    fund_name: str,
    epoch_start: datetime,
    epoch_end: datetime,
    performance_rate: Decimal,
    start_balance: Decimal,
    deposits: Decimal,
    withdrawals: Decimal,
    profit: Decimal,
    end_balance: Decimal,
    previous_hash: str,
) -> str:
    # Stable, explicit field ordering (no locale, no float)
    return "|".join(
        [
            internal_client_code,
            fund_name.lower(),
            epoch_start.isoformat(),
            epoch_end.isoformat(),
            f"{performance_rate:.8f}",
            f"{start_balance:.2f}",
            f"{deposits:.2f}",
            f"{withdrawals:.2f}",
            f"{profit:.2f}",
            f"{end_balance:.2f}",
            previous_hash,
        ]
    )


def _sha256_hex(payload: str) -> str:
    import hashlib
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


GENESIS_HASH = "0" * 64


def _fund_name_from_core_fund(core_fund: Optional[CoreFund]) -> Optional[str]:
    if not core_fund:
        return None
    return (core_fund.fund_name or "").strip() or None


def _get_run_or_404(run_id: int):
    vr = db.session.query(ValuationRun).filter(ValuationRun.id == run_id).first()
    if not vr:
        return None, (jsonify({"status": 404, "message": "Valuation run not found"}), 404)
    if (vr.status or "").strip().lower() != "committed":
        # Avoid null-ledger errors by hiding non-committed runs
        return None, (jsonify({"status": 409, "message": "Valuation run is not committed"}), 409)
    return vr, None


def _run_ledger_aggregates(fund_name: str, epoch_start: datetime, epoch_end: datetime):
    opening_balance_expr = case(
        (
            or_(EpochLedger.start_balance == 0, EpochLedger.start_balance.is_(None)),
            func.coalesce(EpochLedger.deposits, 0)
        ),
        else_=func.coalesce(EpochLedger.start_balance, 0),
    )

    sums = (
        db.session.query(
            func.coalesce(func.sum(opening_balance_expr), 0),
            func.coalesce(func.sum(EpochLedger.deposits), 0),
            func.coalesce(func.sum(EpochLedger.withdrawals), 0),
            func.coalesce(func.sum(EpochLedger.profit), 0),
            func.coalesce(func.sum(EpochLedger.end_balance), 0),
            func.count(EpochLedger.id),
        )
        .filter(
            and_(
                EpochLedger.fund_name == fund_name,
                EpochLedger.epoch_start == epoch_start,
                EpochLedger.epoch_end == epoch_end,
            )
        )
        .first()
    )

    total_opening, total_deposits, total_withdrawals, total_profit, total_closing, investor_count = sums
    return {
        "total_opening_capital": float_2dp(total_opening),
        "total_deposits": float_2dp(total_deposits),
        "total_withdrawals": float_2dp(total_withdrawals),
        "total_profit_distributed": float_2dp(total_profit),
        "total_closing_aum": float_2dp(total_closing),
        "investor_count": int(investor_count or 0),
    }


def _latest_epoch_balances_subquery(as_of: Optional[datetime] = None):
    """
    Helper for portfolio/batch aggregation:
    returns a subquery of the latest EpochLedger row per (internal_client_code, fund_name)
    at or before `as_of` (if provided).
    """
    base = db.session.query(
        EpochLedger.internal_client_code.label("code"),
        EpochLedger.fund_name.label("fund_name"),
        func.max(EpochLedger.epoch_end).label("last_epoch_end"),
    )
    if as_of is not None:
        base = base.filter(EpochLedger.epoch_end <= as_of)
    lep = base.group_by(EpochLedger.internal_client_code, EpochLedger.fund_name).subquery("lep")

    latest = (
        db.session.query(
            EpochLedger.internal_client_code.label("code"),
            EpochLedger.fund_name.label("fund_name"),
            EpochLedger.epoch_start.label("epoch_start"),
            EpochLedger.epoch_end.label("epoch_end"),
            EpochLedger.start_balance.label("start_balance"),
            EpochLedger.deposits.label("deposits"),
            EpochLedger.withdrawals.label("withdrawals"),
            EpochLedger.profit.label("profit"),
            EpochLedger.end_balance.label("end_balance"),
        )
        .join(
            lep,
            and_(
                lep.c.code == EpochLedger.internal_client_code,
                lep.c.fund_name == EpochLedger.fund_name,
                lep.c.last_epoch_end == EpochLedger.epoch_end,
            ),
        )
    ).subquery("latest_epochs")

    return latest


# New canonical endpoints (requested):
# - GET /api/v1/reports
# - GET /api/v1/reports/<valuation_run_id>
# - GET /api/v1/reports/<valuation_run_id>/pdf
# - GET /api/v1/reports/portfolio
# - GET /api/v1/reports/portfolio/multi-batch
# - GET /api/v1/reports/batch/<id>/summary
# - GET /api/v1/reports/batch/<id>/reconciliation


@reports_v1.route("/api/v1/reports/verify-hash-chain/<investor_code>/<fund_name>", methods=["GET"])
@jwt_required()
def verify_hash_chain(investor_code, fund_name):
    """
    Verify the cryptographic integrity of an investor's ledger hash chain.
    """
    try:
        result = verify_ledger_hash_chain(investor_code, fund_name)
        status_code = 200 if result["valid"] else 409
        return jsonify({"status": status_code, "message": result["message"], "data": result}), status_code
    except Exception as e:
        return jsonify({"status": 500, "message": f"Error: {str(e)}"}), 500
@reports_v1.route("/api/v1/reports", methods=["GET"], strict_slashes=False)
@reports_v1.route("/api/v1/reports/", methods=["GET"], strict_slashes=False)
@jwt_required()
def list_reports():
    """
    List committed valuation runs as report rows.
    
    CRITICAL: Only includes ValuationRuns up to the latest commit (e.g., Sept 2026).
    This prevents unprocessed months (like October in 'Principal Only' state) from showing in charts.

    Query Parameters:
    - fund_id: optional CoreFund.id filter
    - limit: optional limit (default: 100)
    """
    try:
        fund_id = request.args.get("fund_id", type=int)
        limit = request.args.get("limit", default=100, type=int)

        # ── CRITICAL: Find the LAST PROCESSED epoch (Sept 2026) ──
        max_chart_epoch = db.session.query(
            func.max(ValuationRun.epoch_end)
        ).filter(
            func.lower(ValuationRun.status) == "committed"
        ).scalar()

        query = db.session.query(ValuationRun).filter(
            func.lower(ValuationRun.status) == "committed"
        )
        
        # Restrict to max_chart_epoch (exclude unprocessed months)
        if max_chart_epoch:
            query = query.filter(ValuationRun.epoch_end <= max_chart_epoch)

        query = query.order_by(desc(ValuationRun.epoch_end))

        if fund_id:
            query = query.filter(ValuationRun.core_fund_id == fund_id)

        valuation_runs = query.limit(limit).all()

        result = []
        for vr in valuation_runs:
            fund = db.session.query(CoreFund).filter(CoreFund.id == vr.core_fund_id).first()
            fund_name = _fund_name_from_core_fund(fund)
            if not fund_name:
                fund_name = f"Unknown fund ({vr.core_fund_id})"

            summary = _run_ledger_aggregates(fund_name, vr.epoch_start, vr.epoch_end)

            result.append({
                "id": vr.id,
                "fund_id": vr.core_fund_id,
                "fund_name": fund_name,
                "epoch_start": vr.epoch_start.isoformat(),
                "epoch_end": vr.epoch_end.isoformat(),
                "performance_rate_percent": float_2dp(float(vr.performance_rate) * 100),
                "head_office_total": float_2dp(vr.head_office_total),
                "summary": summary,
                "status": vr.status,
                "created_at": vr.created_at.isoformat(),
            })

        return jsonify({"status": 200, "message": "Reports retrieved", "data": result}), 200

    except Exception as e:
        return jsonify({"status": 500, "message": f"Error: {str(e)}"}), 500


@reports_v1.route("/api/v1/reports/portfolio", methods=["GET"])
@jwt_required()
def portfolio_view():
    """
    Global portfolio AUM per active core fund at an as-of date.

    Query:
      - as_of (optional ISO date/datetime). Defaults to today (UTC date).
    """
    try:
        as_of_raw = (request.args.get("as_of") or "").strip()
        if as_of_raw:
            try:
                as_of = datetime.fromisoformat(as_of_raw.replace("Z", "+00:00"))
            except Exception:
                return jsonify({"status": 400, "message": "Invalid as_of date"}), 400
        else:
            as_of = datetime.now(timezone.utc)

        latest = _latest_epoch_balances_subquery(as_of)

        rows = (
            db.session.query(
                CoreFund.id.label("core_fund_id"),
                CoreFund.fund_name.label("core_fund_name"),
                func.coalesce(func.sum(latest.c.end_balance), 0).label("total_aum"),
            )
            .outerjoin(
                latest,
                func.lower(latest.c.fund_name) == func.lower(CoreFund.fund_name),
            )
            .filter(CoreFund.is_active.is_(True))
            .group_by(CoreFund.id, CoreFund.fund_name)
            .order_by(CoreFund.fund_name.asc())
            .all()
        )

        data = [
            {
                "core_fund_id": r.core_fund_id,
                "core_fund_name": r.core_fund_name,
                "total_aum": float_2dp(r.total_aum or 0),
            }
            for r in rows
        ]
        total_global = float_2dp(sum(item["total_aum"] for item in data))

        return jsonify(
            {
                "status": 200,
                "message": "Portfolio AUM retrieved",
                "as_of": as_of.isoformat(),
                "total_global_aum": total_global,
                "funds": data,
            }
        ), 200
    except Exception as e:
        return jsonify({"status": 500, "message": f"Error: {str(e)}"}), 500


def _batch_portfolio_rows(batch_ids, fund_filter: Optional[str] = None):
    """
    Aggregation helper for batch-level portfolio summaries.
    Uses the latest EpochLedger row per investor + fund so totals are reconciled
    against the current investor standings rather than summing all ledger history.
    """
    if not batch_ids:
        return []

    # Build a subquery that selects the latest committed ledger row for each investor/fund.
    latest_epoch_subquery = (
        db.session.query(
            EpochLedger.internal_client_code.label("internal_client_code"),
            func.lower(EpochLedger.fund_name).label("fund_name_lower"),
            func.max(EpochLedger.epoch_end).label("latest_epoch_end"),
        )
        .group_by(EpochLedger.internal_client_code, func.lower(EpochLedger.fund_name))
        .subquery()
    )

    latest_epoch = (
        db.session.query(
            EpochLedger.internal_client_code.label("internal_client_code"),
            EpochLedger.fund_name.label("fund_name"),
            EpochLedger.epoch_start.label("epoch_start"),
            EpochLedger.epoch_end.label("epoch_end"),
            EpochLedger.start_balance.label("start_balance"),
            EpochLedger.deposits.label("deposits"),
            EpochLedger.withdrawals.label("withdrawals"),
            EpochLedger.profit.label("profit"),
            EpochLedger.end_balance.label("end_balance"),
        )
        .join(
            latest_epoch_subquery,
            (EpochLedger.internal_client_code == latest_epoch_subquery.c.internal_client_code) &
            (func.lower(EpochLedger.fund_name) == latest_epoch_subquery.c.fund_name_lower) &
            (EpochLedger.epoch_end == latest_epoch_subquery.c.latest_epoch_end)
        )
        .subquery()
    )

    # Direct query approach using the latest ledger row for each investor/fund
    q = (
        db.session.query(
            Batch.id.label("batch_id"),
            Batch.batch_name.label("batch_name"),
            CoreFund.fund_name.label("core_fund_name"),
            func.count(func.distinct(Investment.internal_client_code)).label("investors_count"),
            func.max(latest_epoch.c.epoch_end).label("as_of_epoch_end"),
            func.coalesce(func.sum(latest_epoch.c.start_balance), 0).label("total_opening_capital"),
            func.coalesce(func.sum(latest_epoch.c.deposits), 0).label("total_deposits"),
            func.coalesce(func.sum(latest_epoch.c.withdrawals), 0).label("total_withdrawals"),
            func.coalesce(func.sum(latest_epoch.c.profit), 0).label("total_profit"),
            func.coalesce(func.sum(latest_epoch.c.end_balance), 0).label("total_closing_aum"),
        )
        .join(Investment, Investment.batch_id == Batch.id)
        .join(CoreFund, Investment.fund_id == CoreFund.id)
        .outerjoin(
            latest_epoch,
            (Investment.internal_client_code == latest_epoch.c.internal_client_code) &
            (func.lower(CoreFund.fund_name) == func.lower(latest_epoch.c.fund_name))
        )
        .filter(Batch.id.in_(batch_ids))
    )

    if fund_filter:
        q = q.filter(func.lower(CoreFund.fund_name) == fund_filter.lower())

    rows = (
        q.group_by(Batch.id, Batch.batch_name, CoreFund.fund_name)
        .order_by(Batch.id.asc(), CoreFund.fund_name.asc())
        .all()
    )

    # Attach performance rate from ValuationRun for that "as_of" period (if committed)
    run_keys = {(r.core_fund_name.lower(), r.as_of_epoch_end) for r in rows if r.as_of_epoch_end and r.core_fund_name}
    performance_by_key = {}
    if run_keys:
        # Resolve core_fund_id for names
        names = sorted({name for (name, _) in run_keys})
        core_rows = db.session.query(CoreFund).filter(func.lower(CoreFund.fund_name).in_(names)).all()
        core_by_lower = {c.fund_name.lower(): c.id for c in core_rows}

        # Fetch all matching runs
        core_ids = list(core_by_lower.values())
        vr_rows = db.session.query(ValuationRun).filter(
            ValuationRun.core_fund_id.in_(core_ids),
            ValuationRun.status == "Committed",
        ).all()
        vr_by_key = {(vr.core_fund_id, vr.epoch_end): vr for vr in vr_rows}

        for (name_lower, epoch_end) in run_keys:
            core_id = core_by_lower.get(name_lower)
            if not core_id:
                continue
            vr = vr_by_key.get((core_id, epoch_end))
            if vr:
                performance_by_key[(name_lower, epoch_end)] = float_2dp(float(vr.performance_rate) * 100)

    out = []
    for r in rows:
        key = (r.core_fund_name.lower(), r.as_of_epoch_end) if r.core_fund_name and r.as_of_epoch_end else None
        out.append({
            "batch_id": r.batch_id,
            "batch_name": r.batch_name,
            "core_fund_name": r.core_fund_name,
            "as_of_epoch_end": r.as_of_epoch_end.isoformat() if r.as_of_epoch_end else None,
            "performance_rate_percent": performance_by_key.get(key) if key else None,
            "investors_count": int(r.investors_count or 0),
            "total_opening_capital": float_2dp(r.total_opening_capital or 0),
            "total_deposits": float_2dp(r.total_deposits or 0),
            "total_withdrawals": float_2dp(r.total_withdrawals or 0),
            "total_profit": float_2dp(r.total_profit or 0),
            "total_closing_aum": float_2dp(r.total_closing_aum or 0),
        })
    return out


@reports_v1.route("/api/v1/reports/portfolio/multi-batch", methods=["GET"])
@jwt_required()
def multi_batch_portfolio_excel():
    """
    Legacy endpoint used by GlobalReports "Generate Portfolio Report (Excel)".

    Query:
      - batch_ids: comma-separated list of batch IDs
      - fund_filter (optional): Core fund filter

    Returns:
      - Excel (.xlsx) file with per-batch, per-fund aggregates.
    """
    try:
        batch_ids_raw = (request.args.get("batch_ids") or "").strip()
        batch_ids = []
        if not batch_ids_raw:
            # If no batch_ids provided, default to all active batches
            batch_ids = [b[0] for b in db.session.query(Batch.id).filter(Batch.is_active.is_(True)).all()]
        else:
            # Accept comma-separated list; tolerate tokens like "1:1" by taking the first numeric segment.
            for token in batch_ids_raw.split(","):
                t = (token or "").strip()
                if not t:
                    continue
                if ":" in t:
                    t = t.split(":", 1)[0].strip()
                # keep only digits if someone passes "batch-1"
                digits = "".join(ch for ch in t if ch.isdigit())
                if not digits:
                    continue
                batch_ids.append(int(digits))
            batch_ids = sorted(set(batch_ids))

        if not batch_ids:
            return jsonify({"status": 404, "message": "No active batches found to report on"}), 404

        fund_filter = (request.args.get("fund_name") or "").strip() or None

        rows = _batch_portfolio_rows(batch_ids, fund_filter)
        df = pd.DataFrame(rows or [])

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Portfolio")
        output.seek(0)

        filename = f"Portfolio_Report_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"status": 500, "message": f"Error: {str(e)}"}), 500


@reports_v1.route("/api/v1/reports/batch/<int:batch_id>/summary", methods=["GET"])
@jwt_required()
def get_batch_summary_json(batch_id: int):
    """
    Get batch summary data as JSON (for display in UI).
    Aggregates all investor data by fund for the batch using latest EpochLedger entries.
    Supports COMPOUND GROWTH: opening_balance uses previous epoch's end_balance.
    """
    try:
        batch = db.session.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            return jsonify({"status": 404, "message": "Batch not found"}), 404

        # Get all investments for this batch to build summary
        investments = db.session.query(Investment).filter(Investment.batch_id == batch_id).all()
        
        if not investments:
            return jsonify({"status": 404, "message": "No investments found for this batch"}), 404

        # Group investors by fund to build fund summary
        fund_summary = {}
        total_investors = set()
        total_transactions = 0
        total_opening = Decimal("0")
        total_performance = Decimal("0")
        total_closing = Decimal("0")

        # STEP 1: Get latest epoch ledger entries for each investor/fund
        # This ensures we capture ALL compounding from previous epochs
        for inv in investments:
            fund_name = inv.fund.fund_name if inv.fund else "Unknown"
            investor_email = inv.investor_email
            
            if fund_name not in fund_summary:
                fund_summary[fund_name] = {
                    "investors": set(),
                    "transactions": 0,
                    "opening_balance": Decimal("0"),
                    "performance": Decimal("0"),
                    "closing_total": Decimal("0"),
                }
            
            fund_summary[fund_name]["investors"].add(investor_email)
            fund_summary[fund_name]["transactions"] += 1
            
            total_investors.add(investor_email)
            total_transactions += 1
            
            # Query the MOST RECENT epoch ledger entry for this investor/fund
            # This automatically includes all compounding from previous epochs
            latest_epoch = (
                db.session.query(EpochLedger)
                .filter(
                    EpochLedger.internal_client_code == inv.internal_client_code,
                    func.lower(EpochLedger.fund_name) == fund_name.lower()
                )
                .order_by(EpochLedger.epoch_end.desc())
                .first()
            )
            
            if latest_epoch:
                # Use EpochLedger.start_balance which includes compounding from previous epochs
                opening = Decimal(str(latest_epoch.start_balance or 0))
                profit = Decimal(str(latest_epoch.profit or 0))
                closing = Decimal(str(latest_epoch.end_balance or 0))
                
                fund_summary[fund_name]["opening_balance"] += opening
                fund_summary[fund_name]["performance"] += profit
                fund_summary[fund_name]["closing_total"] += closing
                
                total_opening += opening
                total_performance += profit
                total_closing += closing

        # Convert sets to counts and decimals to floats
        for fund in fund_summary:
            fund_summary[fund]["investors"] = len(fund_summary[fund]["investors"])
            fund_summary[fund]["opening_balance"] = float_2dp(fund_summary[fund]["opening_balance"])
            fund_summary[fund]["performance"] = float_2dp(fund_summary[fund]["performance"])
            fund_summary[fund]["closing_total"] = float_2dp(fund_summary[fund]["closing_total"])

        return jsonify({
            "status": 200,
            "message": "Batch summary retrieved",
            "data": {
                "batch_id": batch_id,
                "batch_name": batch.batch_name,
                "total_investors": len(total_investors),
                "total_transactions": total_transactions,
                "total_opening": float_2dp(total_opening),
                "total_performance": float_2dp(total_performance),
                "total_closing": float_2dp(total_closing),
                "fund_summary": fund_summary
            }
        }), 200
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"status": 500, "message": f"Error: {str(e)}"}), 500


@reports_v1.route("/api/v1/reports/batch/<int:batch_id>/summary-excel", methods=["GET"])
@jwt_required()
def batch_summary_excel(batch_id: int):
    """
    Download batch summary as Excel file with individual investor rows.
    Uses exact investment-level data from the Batch Detail page display.
    """
    try:
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Get batch
        batch = db.session.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            return jsonify({"status": 404, "message": "Batch not found"}), 404
        
        # Get investments for this batch
        investments = db.session.query(Investment).filter(
            Investment.batch_id == batch_id
        ).all()
        
        if not investments:
            # Return empty excel with headers
            df = pd.DataFrame(columns=[
                "Investor Name",
                "Client Code",
                "Fund Name",
                "Email",
                "Initial Batch Principal",
                "Batch Withdrawals",
                "Monthly Profit ($)",
                "Current Batch Standing"
            ])
        else:
            # Get calculated values for each investment
            from app.Batch.controllers import BatchController
            
            rows = []
            total_principal = 0.0
            total_withdrawals = 0.0
            total_profit = 0.0
            total_standing = 0.0
            
            for inv in investments:
                # Get calculated values (includes withdrawal fix)
                epoch_data = BatchController._calculate_batch_investment_values(inv, batch, db.session)
                
                opening_balance = epoch_data.get("opening_balance", 0.0)
                current_balance = epoch_data.get("current_balance", 0.0)
                withdrawals = epoch_data.get("withdrawals", 0.0)
                profit = epoch_data.get("profit", 0.0)
                
                # Build row with exact investor data
                row = {
                    "Investor Name": inv.investor_name or "Unknown",
                    "Client Code": inv.internal_client_code or "—",
                    "Fund Name": inv.fund.fund_name if inv.fund else inv.fund_name or "—",
                    "Email": inv.investor_email or "—",
                    "Initial Batch Principal": float_2dp(opening_balance),
                    "Batch Withdrawals": float_2dp(withdrawals),  # Always show as positive number
                    "Monthly Profit ($)": float_2dp(profit),
                    "Current Batch Standing": float_2dp(current_balance),
                }
                
                rows.append(row)
                
                # Accumulate totals
                total_principal += opening_balance
                total_withdrawals += withdrawals
                total_profit += profit
                total_standing += current_balance
            
            # Add grand total row
            rows.append({
                "Investor Name": "GRAND TOTAL",
                "Client Code": "",
                "Fund Name": "",
                "Email": "",
                "Initial Batch Principal": float_2dp(total_principal),
                "Batch Withdrawals": float_2dp(total_withdrawals),
                "Monthly Profit ($)": float_2dp(total_profit),
                "Current Batch Standing": float_2dp(total_standing),
            })
            
            df = pd.DataFrame(rows)
        
        # Create Excel with openpyxl for formatting
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="BatchSummary")
            
            # Access the worksheet
            ws = writer.sheets["BatchSummary"]
            
            # Apply formatting
            header_fill = PatternFill(start_color="00005b", end_color="00005b", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            total_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
            total_font = Font(bold=True, size=11)
            currency_format = "$#,##0.00"
            
            # Format headers (row 1)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Format data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                is_total_row = row[0].value == "GRAND TOTAL"
                
                for col_idx, cell in enumerate(row, start=1):
                    # Set alignment
                    if col_idx in [5, 6, 7, 8]:  # Currency columns
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        if is_total_row:
                            cell.fill = total_fill
                            cell.font = total_font
                        # Format as currency
                        cell.number_format = currency_format
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        if is_total_row:
                            cell.fill = total_fill
                            cell.font = total_font
            
            # Auto-adjust column widths
            for col_idx, column in enumerate(ws.columns, start=1):
                max_length = 0
                column_letter = ws.cell(row=1, column=col_idx).column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Dynamic filename with batch name and date
        filename = f"Batch_{batch.batch_name.replace(' ', '_')}_{batch_id}_Report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"status": 500, "message": f"Error: {str(e)}"}), 500


@reports_v1.route("/api/v1/reports/batch/<int:batch_id>/reconciliation", methods=["GET"])
@jwt_required()
def batch_reconciliation(batch_id: int):
    """
    JSON reconciliation view for a given batch (used by any legacy UI).
    """
    try:
        rows = _batch_portfolio_rows([batch_id], fund_filter=None)
        total_closing = float_2dp(sum(r["total_closing_aum"] for r in rows))
        total_profit = float_2dp(sum(r["total_profit"] for r in rows))

        return jsonify(
            {
                "status": 200,
                "message": "Batch reconciliation retrieved",
                "data": {
                    "batch_id": batch_id,
                    "lines": rows,
                    "total_closing_aum": total_closing,
                    "total_profit": total_profit,
                },
            }
        ), 200
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"status": 500, "message": f"Error: {str(e)}"}), 500


@reports_v1.route("/api/v1/reports/<int:valuation_run_id>", methods=["GET"])
@jwt_required()
def get_report_detail(valuation_run_id):
    """
    Detailed report based on ValuationRun + EpochLedger.
    """
    try:
        vr, err = _get_run_or_404(valuation_run_id)
        if err:
            return err

        fund = db.session.query(CoreFund).filter(CoreFund.id == vr.core_fund_id).first()
        fund_name = _fund_name_from_core_fund(fund)
        if not fund_name:
            return jsonify({"status": 404, "message": "Core fund not found for this valuation run"}), 404

        # Fetch the actual ledger entries for this run
        ledger_entries = db.session.query(EpochLedger).filter(
            func.lower(EpochLedger.fund_name) == fund_name.lower(),
            EpochLedger.epoch_start == vr.epoch_start,
            EpochLedger.epoch_end == vr.epoch_end,
        ).order_by(asc(EpochLedger.internal_client_code)).all()

        if not ledger_entries:
            return jsonify({"status": 404, "message": "No committed ledger entries found for this valuation run"}), 404

        # Perform reconciliation check before generating report
        ledger_total = Decimal(str(sum(e.end_balance for e in ledger_entries)))
        head_office_total = Decimal(str(vr.head_office_total))
        reconciliation_diff = float_2dp(ledger_total - head_office_total)

        # Requirement 5: Reconciliation Guard (Block rendering if mismatch > $0.01)
        if abs(reconciliation_diff) > 0.01:
            print(f"RECONCILIATION MISMATCH: Run {valuation_run_id}, Fund {fund_name}, "
                  f"Ledger: ${ledger_total}, Head Office: ${head_office_total}, Diff: ${reconciliation_diff}")
            
            return jsonify({
                "status": 409,
                "message": f"Reconciliation Mismatch Detected (${reconciliation_diff}). Report distribution blocked for security.",
                "data": {
                    "ledger_total": float_2dp(ledger_total),
                    "head_office_total": float_2dp(head_office_total),
                    "difference": reconciliation_diff
                }
            }), 409

        reconciliation_status = "VERIFIED"
        summary = _run_ledger_aggregates(fund_name, vr.epoch_start, vr.epoch_end)

        # Build investor breakdown
        investor_breakdown = []
        from app.Batch.controllers import BatchController
        
        for entry in ledger_entries:
            inv = (
                db.session.query(Investment)
                .filter(Investment.internal_client_code == entry.internal_client_code)
                .order_by(desc(Investment.id))
                .first()
            )
            # ✅ FIX: Use start_balance directly from ledger (NOT replaced by deposits)
            # start_balance represents opening capital before this period's activity
            # For first period: start_balance = 0 (new investor)
            # For subsequent periods: start_balance = previous epoch's end_balance
            opening_balance = Decimal(str(entry.start_balance or 0))

            # ✅ FIX: Calculate CURRENT balance using same BatchController method as portfolio/directory endpoints
            # This ensures report shows accurate current values, not just historical ledger data
            current_balance = Decimal("0.00")
            try:
                investor_investments = db.session.query(Investment).filter(
                    Investment.internal_client_code == entry.internal_client_code
                ).all()
                
                for inv_item in investor_investments:
                    batch_id = inv_item.batch_id
                    batch = db.session.query(Batch).filter(Batch.id == batch_id).first() if batch_id else None
                    inv_values = BatchController._calculate_batch_investment_values(inv_item, batch, db.session)
                    current_balance += Decimal(str(inv_values["current_balance"]))
            except Exception as e:
                print(f"⚠️  Warning: Balance calculation failed for {entry.internal_client_code}: {str(e)}")
                # Fallback: use ledger end_balance
                current_balance = Decimal(str(entry.end_balance or 0))

            # ✅ FIX: Calculate CURRENT profit based on actual current balance
            # Profit = Current Balance - (Opening Balance + Deposits - Withdrawals)
            deposits_val = Decimal(str(entry.deposits or 0))
            withdrawals_val = Decimal(str(entry.withdrawals or 0))
            current_profit = current_balance - (opening_balance + deposits_val - withdrawals_val)

            investor_breakdown.append({
                "internal_client_code": entry.internal_client_code,
                "investor_name": (inv.investor_name if inv and inv.investor_name else "Unknown"),
                "investor_email": (inv.investor_email if inv and inv.investor_email else None),
                "batch_id": (inv.batch_id if inv and inv.batch_id else None),
                "start_balance": float_2dp(opening_balance),
                "deposits": float_2dp(deposits_val),
                "withdrawals": float_2dp(withdrawals_val),
                "pro_rata_profit": float_2dp(current_profit),
                "end_balance": float_2dp(current_balance),
            })

        # Reconciliation: compare local closing AUM to head_office_total (the commit-time lock value)
        reconciliation_diff = float_2dp(Decimal(str(summary["total_closing_aum"])) - Decimal(str(float_2dp(vr.head_office_total))))

        # Determine linked batch id (if statements are present for this run)
        batch_id = None
        statements_for_run = db.session.query(Statement).filter(Statement.valuation_run_id == valuation_run_id).all()
        if statements_for_run:
            batch_id = statements_for_run[0].batch_id

        return jsonify({
            "status": 200,
            "message": "Report retrieved",
            "data": {
                "id": vr.id,
                "fund_id": vr.core_fund_id,
                "fund_name": fund_name,
                "batch_id": batch_id,
                "epoch_start": vr.epoch_start.isoformat(),
                "epoch_end": vr.epoch_end.isoformat(),
                "performance_rate_percent": float_2dp(float(vr.performance_rate) * 100),
                "head_office_total": float_2dp(vr.head_office_total),
                "status": vr.status,
                "created_at": vr.created_at.isoformat(),
                "reconciliation_status": reconciliation_status,
                "reconciliation_diff": reconciliation_diff,
                "summary": {
                    **summary,
                    "reconciliation": {
                        "status": reconciliation_status,
                        "head_office_total": float_2dp(vr.head_office_total),
                        "local_total_closing_aum": summary["total_closing_aum"],
                        "difference": reconciliation_diff,
                    },
                },
                "investor_breakdown": investor_breakdown,
            }
        }), 200

    except Exception as e:
        return jsonify({"status": 500, "message": f"Error: {str(e)}"}), 500


@reports_v1.route("/api/v1/reports/<int:valuation_run_id>/pdf", methods=["GET"])
@jwt_required()
def get_report_pdf(valuation_run_id):
    """
    Generate PDF report for a specific valuation run (AIB-AXYS branded).
    """
    try:
        vr, err = _get_run_or_404(valuation_run_id)
        if err:
            return err

        fund = db.session.query(CoreFund).filter(CoreFund.id == vr.core_fund_id).first()
        fund_name = _fund_name_from_core_fund(fund)
        if not fund_name:
            return jsonify({"status": 404, "message": "Core fund not found for this valuation run"}), 404

        # Get epoch ledger entries
        ledger_entries = db.session.query(EpochLedger).filter(
            and_(
                EpochLedger.fund_name == fund_name,
                EpochLedger.epoch_start == vr.epoch_start,
                EpochLedger.epoch_end == vr.epoch_end,
            )
        ).order_by(asc(EpochLedger.internal_client_code)).all()

        # Calculate summary
        # ✅ FIX: Use start_balance directly, not "or entry.deposits"
        # start_balance already accounts for all prior periods + deposits
        total_start_balance = sum(float(entry.start_balance or 0) for entry in ledger_entries)
        total_deposits = sum(float(entry.deposits) for entry in ledger_entries)
        total_withdrawals = sum(float(entry.withdrawals) for entry in ledger_entries)
        total_profit = sum(float(entry.profit) for entry in ledger_entries)
        total_end_balance = sum(float(entry.end_balance) for entry in ledger_entries)
        
        # Create PDF
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#00005b'),
            spaceAfter=6,
            fontName='Helvetica-Bold'
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#00005b'),
            spaceAfter=12,
            fontName='Helvetica-Bold'
        )
        
        # Header
        story.append(Paragraph("AIB-AXYS Africa", title_style))
        story.append(Paragraph("Valuation Period Report", heading_style))
        story.append(Spacer(1, 0.1*inch))
        
        # Report Info
        period_start = vr.epoch_start.strftime('%B %d, %Y')
        period_end = vr.epoch_end.strftime('%B %d, %Y')
        
        report_data = [
            ['Fund Name:', fund_name],
            ['Performance Period:', f'{period_start} to {period_end}'],
            ['Performance Rate:', f'{float_2dp(float(vr.performance_rate) * 100)}%'],
            ['Report Generated:', datetime.now().strftime('%B %d, %Y')],
        ]
        report_table = Table(report_data, colWidths=[1.5*inch, 3*inch])
        report_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#00005b')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(report_table)
        story.append(Spacer(1, 0.2*inch))
        
        # Financial Summary
        story.append(Paragraph("Financial Summary", heading_style))
        
        # Add note about compound growth
        summary_note = "Opening Capital includes compound growth from all previous epochs. " \
                      "Profit is allocated pro-rata based on weighted capital share."
        story.append(Paragraph(summary_note, styles['Normal']))
        story.append(Spacer(1, 0.08*inch))
        
        summary_data = [
            ['Opening Capital', format_currency(total_start_balance)],
            ['Total Deposits', format_currency(total_deposits)],
            ['Total Withdrawals', f'({format_currency(total_withdrawals)})'],
            ['Total Profit', format_currency(total_profit)],
            ['Closing AUM', format_currency(total_end_balance)],
            ['Head Office Total', format_currency(vr.head_office_total)],
        ]
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 11),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#00005b')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 4), (-1, 5), colors.HexColor('#e6e6f0')),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.2*inch))
        
        # Investor Breakdown
        if ledger_entries:
            story.append(Paragraph("Investor Breakdown (Compound Growth Progression)", heading_style))
            
            # Add a note about compound growth
            compound_note = "Note: Starting balance includes compound growth from previous epochs. " \
                           "Profit is calculated on (Start Balance + Deposits - Withdrawals)."
            story.append(Paragraph(compound_note, styles['Normal']))
            story.append(Spacer(1, 0.1*inch))
            
            breakdown_data = [
                ['Investor Code', 'Name', 'Start Balance', 'Deposits', 'Withdrawals', 'Profit', 'End Balance']
            ]
            
            # ✅ FIX: Use same BatchController calculation as web report
            from app.Batch.controllers import BatchController
            
            for entry in ledger_entries:
                investments = db.session.query(Investment).filter(
                    Investment.internal_client_code == entry.internal_client_code
                ).all()
                investor_name = investments[0].investor_name if investments else "Unknown"
                opening_balance = Decimal(str(entry.start_balance or 0))
                
                # Calculate CURRENT balance using BatchController (same as web report)
                current_balance = Decimal("0.00")
                try:
                    for inv_item in investments:
                        batch_id = inv_item.batch_id
                        batch = db.session.query(Batch).filter(Batch.id == batch_id).first() if batch_id else None
                        inv_values = BatchController._calculate_batch_investment_values(inv_item, batch, db.session)
                        current_balance += Decimal(str(inv_values["current_balance"]))
                except Exception as e:
                    print(f"⚠️  Warning: Balance calculation failed for {entry.internal_client_code}: {str(e)}")
                    current_balance = Decimal(str(entry.end_balance or 0))
                
                # Calculate CURRENT profit
                deposits_val = Decimal(str(entry.deposits or 0))
                withdrawals_val = Decimal(str(entry.withdrawals or 0))
                current_profit = current_balance - (opening_balance + deposits_val - withdrawals_val)
                
                breakdown_data.append([
                    entry.internal_client_code,
                    investor_name[:20],  # Truncate long names
                    format_currency(float_2dp(opening_balance)),
                    format_currency(float_2dp(deposits_val)),
                    format_currency(float_2dp(withdrawals_val)),
                    format_currency(float_2dp(current_profit)),
                    format_currency(float_2dp(current_balance)),
                ])
            
            breakdown_table = Table(breakdown_data, colWidths=[0.85*inch, 1.3*inch, 0.95*inch, 0.85*inch, 0.95*inch, 0.85*inch, 1*inch])
            breakdown_table.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, -1), 'Helvetica', 8),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00005b')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7fa')]),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(breakdown_table)
        
        # Build PDF
        doc.build(story)
        pdf_buffer.seek(0)
        
        filename = f"Valuation_{fund_name}_{vr.epoch_end.strftime('%Y%m%d')}.pdf"
        
        return send_file(
            pdf_buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({"status": 500, "message": f"Error: {str(e)}"}), 500


# Backwards-compatible endpoints used by existing frontend service (keep for now)
@reports_v1.route("/api/v1/reports/valuation-runs", methods=["GET"])
@jwt_required()
def list_valuation_runs_compat():
    return list_reports()


@reports_v1.route("/api/v1/reports/valuation-runs/<int:valuation_run_id>", methods=["GET"])
@jwt_required()
def get_valuation_run_detail_compat(valuation_run_id):
    return get_report_detail(valuation_run_id)


@reports_v1.route("/api/v1/reports/valuation-runs/<int:valuation_run_id>/pdf", methods=["GET"])
@jwt_required()
def get_valuation_run_pdf_compat(valuation_run_id):
    return get_report_pdf(valuation_run_id)


@reports_v1.route("/api/v1/reports/batch-summary/<int:batch_id>", methods=["GET"])
@jwt_required()
def get_batch_summary_statement(batch_id):
    """
    Report Type 1: Batch Summary Statement (Internal)
    Summary for the entire Batch, grouped by Fund.
    """
    try:
        batch = db.session.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            return jsonify({"status": 404, "message": "Batch not found"}), 404

        # Get statements for this batch
        statements = db.session.query(Statement).filter(Statement.batch_id == batch_id).all()

        if not statements:
            return jsonify({"status": 404, "message": "No statements found for this batch"}), 404

        # Group by fund
        fund_summary = {}
        total_investors = set()
        total_transactions = 0
        total_performance = Decimal("0")

        for stmt in statements:
            fund_name = stmt.fund.fund_name if stmt.fund else "Unknown"
            if fund_name not in fund_summary:
                fund_summary[fund_name] = {
                    "investors": set(),
                    "transactions": 0,
                    "performance": Decimal("0"),
                    "closing_total": Decimal("0")
                }
            investor_code = stmt.investor.internal_client_code if stmt.investor and stmt.investor.internal_client_code else stmt.investor.investor_email
            fund_summary[fund_name]["investors"].add(investor_code)
            fund_summary[fund_name]["transactions"] += 1
            fund_summary[fund_name]["performance"] += stmt.performance_gain
            fund_summary[fund_name]["closing_total"] += stmt.closing_balance
            total_investors.add(investor_code)
            total_transactions += 1
            total_performance += stmt.performance_gain

        # Convert sets to counts
        for fund in fund_summary:
            fund_summary[fund]["investors"] = len(fund_summary[fund]["investors"])

        return jsonify({
            "status": 200,
            "message": "Batch summary statement retrieved",
            "data": {
                "batch_id": batch_id,
                "batch_name": batch.batch_name,
                "total_investors": len(total_investors),
                "total_transactions": total_transactions,
                "total_performance": float(total_performance),
                "fund_summary": fund_summary
            }
        }), 200
    except Exception as e:
        return jsonify({"status": 500, "message": f"Error: {str(e)}"}), 500


@reports_v1.route("/api/v1/reports/investor-statement/<int:batch_id>/<investor_email>", methods=["GET"])
@jwt_required()
def get_investor_statement(batch_id, investor_email):
    """
    Report Type 2: Individual Investor Statement (Client Facing)
    For a single investor in a specific batch, show all their fund holdings and transactions.
    Generates statement dynamically from EpochLedger and Investment data.
    """
    try:
        batch = db.session.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            return jsonify({"status": 404, "message": "Batch not found"}), 404

        # Get investor investment record to verify existence
        investor_inv = (
            db.session.query(Investment)
            .filter(
                Investment.batch_id == batch_id,
                Investment.investor_email == investor_email
            )
            .first()
        )

        if not investor_inv:
            return jsonify({"status": 404, "message": "No investor found in this batch"}), 404

        investor_name = investor_inv.investor_name if investor_inv else investor_email
        
        # Determine the statement epoch scope (by valuation_run_id or explicit date range)
        valuation_run_id = request.args.get("valuation_run_id", type=int)
        epoch_start = None
        epoch_end = None

        if valuation_run_id:
            vr = db.session.query(ValuationRun).filter(ValuationRun.id == valuation_run_id).first()
            if not vr:
                return jsonify({"status": 404, "message": "Valuation run not found"}), 404
            epoch_start = vr.epoch_start
            epoch_end = vr.epoch_end
        else:
            epoch_start_str = request.args.get("epoch_start")
            epoch_end_str = request.args.get("epoch_end")
            if not epoch_start_str or not epoch_end_str:
                return jsonify({"status": 400, "message": "valuation_run_id or epoch_start and epoch_end required"}), 400
            try:
                epoch_start = datetime.fromisoformat(epoch_start_str)
                epoch_end = datetime.fromisoformat(epoch_end_str)
            except Exception:
                return jsonify({"status": 400, "message": "Invalid epoch_start/epoch_end format"}), 400

        # Get epoch ledger entries for this investor in the requested period
        epoch_entries = (
            db.session.query(EpochLedger)
            .filter(
                EpochLedger.internal_client_code == investor_inv.internal_client_code,
                EpochLedger.epoch_start == epoch_start,
                EpochLedger.epoch_end == epoch_end,
            )
            .order_by(EpochLedger.fund_name.asc())
            .all()
        )

        if not epoch_entries:
            return jsonify({"status": 404, "message": "No epoch ledger entries found for this investor in the requested period"}), 404

        # Build fund holdings from the scoped epoch ledger entries (by epoch_start/epoch_end)
        investor_data = {}
        for entry in epoch_entries:
            # ✅ FIX: Use start_balance directly, not "or entry.deposits"
            opening_balance = float_2dp(entry.start_balance or 0)
            if entry.fund_name not in investor_data:
                investor_data[entry.fund_name] = {
                    "fund_name": entry.fund_name,
                    "opening_balance": opening_balance,
                    "performance_gain": float_2dp(entry.profit),
                    "closing_balance": float_2dp(entry.end_balance)
                }
            else:
                # If multiple entries in same period/fund, keep the latest end balance for safety
                if entry.end_balance >= Decimal(str(investor_data[entry.fund_name]["closing_balance"])):
                    investor_data[entry.fund_name] = {
                        "fund_name": entry.fund_name,
                        # ✅ FIX: Use start_balance directly, not "or entry.deposits"
                        "opening_balance": float_2dp(entry.start_balance or 0),
                        "performance_gain": float_2dp(entry.profit),
                        "closing_balance": float_2dp(entry.end_balance)
                    }

        funds_list = list(investor_data.values())
        
        # Calculate totals
        total_opening = sum(float(f["opening_balance"]) for f in funds_list)
        total_performance = sum(float(f["performance_gain"]) for f in funds_list)
        total_closing = sum(float(f["closing_balance"]) for f in funds_list)

        return jsonify({
            "status": 200,
            "message": "Investor statement retrieved",
            "data": {
                "batch_id": batch_id,
                "batch_name": batch.batch_name,
                "investor_email": investor_email,
                "investor_name": investor_name,
                "epoch_start": epoch_start.isoformat(),
                "epoch_end": epoch_end.isoformat(),
                "funds": funds_list,
                "summary": {
                    "total_opening_balance": float_2dp(total_opening),
                    "total_performance_gain": float_2dp(total_performance),
                    "total_closing_balance": float_2dp(total_closing)
                }
            }
        }), 200
    except Exception as e:
        return jsonify({"status": 500, "message": f"Error: {str(e)}"}), 500


@reports_v1.route("/api/v1/reports/investor/<investor_email>/pdf", methods=["GET"])
@jwt_required()
def get_investor_statement_pdf(investor_email):
    """
    Generate individual investor statement PDF combining all funds for the investor.
    Uses EpochLedger data to show opening balance, performance gain, and closing balance.
    Branded with AIB-AXYS Africa logo and #00005b styling.
    """
    try:
        # Get batch_id from query parameter
        batch_id = request.args.get("batch_id", type=int)
        if not batch_id:
            return jsonify({"status": 400, "message": "batch_id query parameter required"}), 400

        batch = db.session.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            return jsonify({"status": 404, "message": "Batch not found"}), 404

        # Get investor investment record
        investor_inv = (
            db.session.query(Investment)
            .filter(
                Investment.batch_id == batch_id,
                Investment.investor_email == investor_email
            )
            .first()
        )

        if not investor_inv:
            return jsonify({"status": 404, "message": "No investor found in this batch"}), 404

        investor_name = investor_inv.investor_name if investor_inv else investor_email
        
        # Determine the statement epoch scope (by valuation_run_id or explicit date range)
        valuation_run_id = request.args.get("valuation_run_id", type=int)
        epoch_start = None
        epoch_end = None

        if valuation_run_id:
            vr = db.session.query(ValuationRun).filter(ValuationRun.id == valuation_run_id).first()
            if not vr:
                return jsonify({"status": 404, "message": "Valuation run not found"}), 404
            epoch_start = vr.epoch_start
            epoch_end = vr.epoch_end
        else:
            epoch_start_str = request.args.get("epoch_start")
            epoch_end_str = request.args.get("epoch_end")
            if not epoch_start_str or not epoch_end_str:
                return jsonify({"status": 400, "message": "valuation_run_id or epoch_start and epoch_end required"}), 400
            try:
                epoch_start = datetime.fromisoformat(epoch_start_str)
                epoch_end = datetime.fromisoformat(epoch_end_str)
            except Exception:
                return jsonify({"status": 400, "message": "Invalid epoch_start/epoch_end format"}), 400

        # Get epoch ledger entries for this investor in requested period
        epoch_entries = (
            db.session.query(EpochLedger)
            .filter(
                EpochLedger.internal_client_code == investor_inv.internal_client_code,
                EpochLedger.epoch_start == epoch_start,
                EpochLedger.epoch_end == epoch_end,
            )
            .order_by(EpochLedger.fund_name.asc())
            .all()
        )

        if not epoch_entries:
            return jsonify({"status": 404, "message": "No epoch ledger entries found for this investor in the requested period"}), 404

        # Build fund data from epoch ledger (latest epoch for each fund)
        fund_data_map = {}
        for entry in epoch_entries:
            # ✅ FIX: Use start_balance directly, not "or entry.deposits"
            opening_balance = float_2dp(entry.start_balance or 0)
            if entry.fund_name not in fund_data_map:
                fund_data_map[entry.fund_name] = {
                    "opening_balance": opening_balance,
                    "performance_gain": float_2dp(entry.profit),
                    "closing_balance": float_2dp(entry.end_balance)
                }
            else:
                # Keep latest entry
                if entry.end_balance > Decimal(str(fund_data_map[entry.fund_name]["closing_balance"])):
                    fund_data_map[entry.fund_name] = {
                        # ✅ FIX: Use start_balance directly, not "or entry.deposits"
                        "opening_balance": float_2dp(entry.start_balance or 0),
                        "performance_gain": float_2dp(entry.profit),
                        "closing_balance": float_2dp(entry.end_balance)
                    }

        # Create PDF
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#00005b'),
            spaceAfter=6,
            fontName='Helvetica-Bold'
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#00005b'),
            spaceAfter=12,
            fontName='Helvetica-Bold'
        )
        subheading_style = ParagraphStyle(
            'CustomSubheading',
            parent=styles['Heading3'],
            fontSize=12,
            textColor=colors.HexColor('#00005b'),
            spaceAfter=8,
            fontName='Helvetica-Bold'
        )

        # Header
        story.append(Paragraph("AIB-AXYS Africa", title_style))
        story.append(Paragraph("Investor Statement", heading_style))
        story.append(Spacer(1, 0.1*inch))

        # Investor Info
        period_start = epoch_start.strftime('%B %d, %Y')
        period_end = epoch_end.strftime('%B %d, %Y')

        investor_data = [
            ['Investor Name:', investor_name],
            ['Email:', investor_email],
            ['Batch:', batch.batch_name],
            ['Statement Date:', datetime.now().strftime('%B %d, %Y')],
            ['Period Ending:', period_end],
            ['Period Range:', f'{period_start} to {period_end}'],
        ]
        investor_table = Table(investor_data, colWidths=[1.5*inch, 3*inch])
        investor_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#00005b')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(investor_table)
        story.append(Spacer(1, 0.2*inch))

        # Calculate totals
        total_opening = sum(float(f["opening_balance"]) for f in fund_data_map.values())
        total_performance = sum(float(f["performance_gain"]) for f in fund_data_map.values())
        total_closing = sum(float(f["closing_balance"]) for f in fund_data_map.values())

        story.append(Paragraph("Portfolio Summary", subheading_style))
        summary_data = [
            ['Total Opening Balance', format_currency(total_opening)],
            ['Total Performance Gain', format_currency(total_performance)],
            ['Total Closing Balance', format_currency(total_closing)],
        ]
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 11),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#00005b')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f2f2f2')),
            ('GRID', (0, 0), (-1, -1), 0, colors.transparent),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.2*inch))

        # Fund-by-fund breakdown
        story.append(Paragraph("Fund Breakdown", subheading_style))

        fund_breakdown = [['Fund Name', 'Opening Balance', 'Performance Gain', 'Closing Balance']]
        for fund_name, data in sorted(fund_data_map.items()):
            fund_breakdown.append([
                fund_name,
                format_currency(data["opening_balance"]),
                format_currency(data["performance_gain"]),
                format_currency(data["closing_balance"]),
            ])

        fund_table = Table(fund_breakdown, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        fund_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#00005b')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f2f2f2')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0, colors.transparent),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7fa')]),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(fund_table)

        # Build PDF
        doc.build(story)
        pdf_buffer.seek(0)

        filename = f"Statement_{investor_name.replace(' ', '')}_{batch.batch_name}_{datetime.now().strftime('%Y%m%d')}.pdf"

        return send_file(
            pdf_buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"status": 500, "message": f"Error: {str(e)}"}), 500


@reports_v1.route("/api/v1/reports/batch/<int:batch_id>/summary", methods=["GET"])
@jwt_required()
def get_batch_summary_excel(batch_id):
    """
    Generate Excel summary for a batch (AIB-AXYS branded).
    Includes investor details and fund totals.
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        
        batch = db.session.query(Batch).filter(Batch.id == batch_id).first()
        if not batch:
            return jsonify({"status": 404, "message": "Batch not found"}), 404

        # Get all statements for this batch
        statements = db.session.query(Statement).filter(Statement.batch_id == batch_id).all()
        if not statements:
            return jsonify({"status": 404, "message": "No committed valuation data found for this batch"}), 404

        # Prepare data for Excel
        data = []
        for stmt in statements:
            data.append({
                "Batch ID": batch_id,
                "Investor Name": stmt.investor.investor_name if stmt.investor else "Unknown",
                "Client Code": stmt.investor.internal_client_code if stmt.investor else "N/A",
                "Email": stmt.investor.investor_email if stmt.investor else "N/A",
                "Fund": stmt.fund.fund_name if stmt.fund else "Unknown",
                "Opening Capital ($)": float_2dp(stmt.opening_balance),
                "Performance Profit ($)": float_2dp(stmt.performance_gain),
                "Closing AUM ($)": float_2dp(stmt.closing_balance)
            })

        df = pd.DataFrame(data)

        # Create Excel in memory
        output = BytesIO()
        # Note: 'openpyxl' engine is required for styling
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Batch Summary')
            
            # Apply basic formatting via openpyxl
            workbook = writer.book
            worksheet = writer.sheets['Batch Summary']
            
            header_fill = PatternFill(start_color="00005B", end_color="00005B", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Adjust column width
            for column_cells in worksheet.columns:
                max_len = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_len:
                            max_len = len(str(cell.value))
                    except: pass
                worksheet.column_dimensions[column].width = max_len + 2

        output.seek(0)
        filename = f"Batch_Summary_{batch.batch_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"status": 500, "message": f"Error: {str(e)}"}), 500

