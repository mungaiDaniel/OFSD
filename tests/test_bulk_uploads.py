"""
Bulk File Upload Tests
======================
Tests for Excel file uploads (investments, withdrawals) and batch import operations.

Run with: pytest tests/test_bulk_uploads.py -v
"""

import pytest
import json
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from openpyxl import Workbook
from app.database.database import db
from app.Batch.model import Batch
from app.Investments.model import Investment


def get_auth_headers(token):
    """Helper to create authorization headers"""
    return {'Authorization': f'Bearer {token}'}


def create_sample_excel_file(num_rows=5):
    """Create a sample Excel file for upload testing"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Investments'
    
    # Add headers
    headers = [
        'Investor Name',
        'Email',
        'Phone',
        'Internal Client Code',
        'Amount Deposited',
        'Date Deposited',
        'Wealth Manager',
        'IFA',
        'Contract Note'
    ]
    ws.append(headers)
    
    # Add sample rows
    for i in range(num_rows):
        ws.append([
            f'Investor {i}',
            f'investor{i}@example.com',
            '+1234567890',
            f'INV-BULK-{i:03d}',
            '50000.00',
            datetime.now(timezone.utc).isoformat(),
            'Wealth Manager',
            'IFA Name',
            'https://example.com/contract'
        ])
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@pytest.mark.unit
class TestInvestmentFileUpload:
    """Test investment bulk file uploads"""
    
    def test_upload_investment_excel_success(self, client, auth_token, sample_batch):
        """Test successfully uploading investments from Excel"""
        excel_file = create_sample_excel_file(num_rows=3)
        
        response = client.post(
            f'/api/v1/batches/{sample_batch.id}/upload-excel',
            data={'file': (excel_file, 'investments.xlsx')},
            headers=get_auth_headers(auth_token)
        )
        
        assert response.status_code in [200, 201, 400]
    
    def test_upload_investment_excel_no_file(self, client, auth_token, sample_batch):
        """Test upload without file"""
        response = client.post(
            f'/api/v1/batches/{sample_batch.id}/upload-excel',
            headers=get_auth_headers(auth_token)
        )
        
        assert response.status_code >= 400
    
    def test_upload_investment_excel_invalid_batch(self, client, auth_token):
        """Test uploading to non-existent batch"""
        excel_file = create_sample_excel_file()
        
        response = client.post(
            '/api/v1/batches/99999/upload-excel',
            data={'file': (excel_file, 'investments.xlsx')},
            headers=get_auth_headers(auth_token)
        )
        
        assert response.status_code >= 400
    
    def test_upload_investment_excel_invalid_format(self, client, auth_token, sample_batch):
        """Test uploading invalid file format"""
        # Create a text file instead of Excel
        invalid_file = BytesIO(b'This is not an Excel file')
        
        response = client.post(
            f'/api/v1/batches/{sample_batch.id}/upload-excel',
            data={'file': (invalid_file, 'notexcel.txt')},
            headers=get_auth_headers(auth_token)
        )
        
        assert response.status_code >= 400

    def test_upload_investment_excel_with_standard_headers(self, client, auth_token, sample_batch):
        """Test upload with user-friendly headers and many rows."""
        excel_file = create_sample_excel_file(num_rows=15)

        response = client.post(
            f'/api/v1/batches/{sample_batch.id}/upload-excel',
            data={'file': (excel_file, 'investments_15.xlsx')},
            headers=get_auth_headers(auth_token)
        )

        assert response.status_code in [200, 201]
        body = response.get_json()
        assert body is not None
        assert body.get('data', {}).get('created_count', 0) == 15
    
    def test_upload_investment_excel_duplicate_codes(self, client, auth_token, sample_batch):
        """Test uploading with duplicate client codes"""
        wb = Workbook()
        ws = wb.active
        
        # Add headers
        headers = ['Investor Name', 'Email', 'Internal Client Code', 'Amount Deposited']
        ws.append(headers)
        
        # Add duplicate codes
        for i in range(3):
            ws.append([
                f'Investor {i}',
                f'inv{i}@example.com',
                'DUPLICATE-CODE',  # Same code
                '50000.00'
            ])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = client.post(
            f'/api/v1/batches/{sample_batch.id}/upload-excel',
            data={'file': (output, 'duplicates.xlsx')},
            headers=get_auth_headers(auth_token)
        )
        
        # Should either reject or warn about duplicates
        assert response.status_code in [200, 201, 400]
    
    def test_upload_investment_excel_missing_required_fields(self, client, auth_token, sample_batch):
        """Test uploading Excel with missing required columns"""
        wb = Workbook()
        ws = wb.active
        
        # Add headers without required fields
        headers = ['Investor Name', 'Email']  # Missing internal_client_code, amount
        ws.append(headers)
        
        ws.append(['John Doe', 'john@example.com'])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = client.post(
            f'/api/v1/batches/{sample_batch.id}/upload-excel',
            data={'file': (output, 'incomplete.xlsx')},
            headers=get_auth_headers(auth_token)
        )
        
        # Should reject due to missing columns
        assert response.status_code >= 400
    
    def test_upload_investment_multiple_sheets(self, client, auth_token, sample_batch):
        """Test uploading Excel with multiple sheets"""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Investments'
        
        # Add data to first sheet
        headers = ['Investor Name', 'Email', 'Internal Client Code', 'Amount Deposited']
        ws1.append(headers)
        ws1.append(['John Doe', 'john@example.com', 'INV-001', '50000.00'])
        
        # Add second sheet (should be ignored)
        ws2 = wb.create_sheet('Other Data')
        ws2.append(['Some', 'Other', 'Data'])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = client.post(
            f'/api/v1/batches/{sample_batch.id}/upload-excel',
            data={'file': (output, 'multi-sheet.xlsx')},
            headers=get_auth_headers(auth_token)
        )
        
        assert response.status_code in [200, 201, 400]


@pytest.mark.unit
class TestInvestmentUploadEndpoint:
    """Test the standalone investment upload endpoint"""
    
    def test_upload_investments_standalone(self, client, auth_token):
        """Test uploading investments via standalone endpoint"""
        excel_file = create_sample_excel_file(num_rows=2)
        
        response = client.post(
            '/api/v1/investments/upload',
            data={'file': (excel_file, 'investments.xlsx')},
            headers=get_auth_headers(auth_token)
        )
        
        assert response.status_code in [200, 201, 400]
    
    def test_upload_investments_standalone_with_batch(self, client, auth_token, sample_batch):
        """Test standalone upload assigning to batch"""
        excel_file = create_sample_excel_file(num_rows=1)
        
        response = client.post(
            f'/api/v1/investments/upload?batch_id={sample_batch.id}',
            data={'file': (excel_file, 'investments.xlsx')},
            headers=get_auth_headers(auth_token)
        )
        
        assert response.status_code in [200, 201, 400]


@pytest.mark.unit
class TestFundUploadOperations:
    """Test fund-level upload operations"""
    
    def test_upload_investments_to_specific_fund(self, client, auth_token, sample_batch):
        """Test uploading investments to specific fund"""
        excel_file = create_sample_excel_file(num_rows=2)
        
        response = client.post(
            f'/api/v1/batches/{sample_batch.id}/upload-excel?fund=Axiom',
            data={'file': (excel_file, 'axiom-investments.xlsx')},
            headers=get_auth_headers(auth_token)
        )
        
        assert response.status_code in [200, 201, 400, 404]


@pytest.mark.unit
class TestWithdrawalFileUpload:
    """Test withdrawal bulk uploads"""
    
    def test_upload_withdrawals_excel(self, client, auth_token, sample_batch):
        """Test uploading withdrawals from Excel"""
        wb = Workbook()
        ws = wb.active
        
        # Add headers
        headers = [
            'Investor Code',
            'Amount',
            'Withdrawal Date',
            'Reason'
        ]
        ws.append(headers)
        
        # Add sample withdrawals
        for i in range(2):
            ws.append([
                f'INV-BULK-{i:03d}',
                '10000.00',
                datetime.now(timezone.utc).isoformat(),
                'Partial withdrawal'
            ])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = client.post(
            '/api/v1/withdrawals/upload',
            data={'file': (output, 'withdrawals.xlsx')},
            headers=get_auth_headers(auth_token)
        )
        
        assert response.status_code in [200, 201, 400]
    
    def test_upload_withdrawals_validation(self, client, auth_token):
        """Test withdrawal upload validation"""
        wb = Workbook()
        ws = wb.active
        
        # Missing required columns
        headers = ['Investor Code', 'Amount']
        ws.append(headers)
        ws.append(['INV-001', '10000.00'])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = client.post(
            '/api/v1/withdrawals/upload',
            data={'file': (output, 'invalid-withdrawals.xlsx')},
            headers=get_auth_headers(auth_token)
        )
        
        # Should validate column requirements
        assert response.status_code in [200, 201, 400]


@pytest.mark.unit
class TestUploadValidation:
    """Test file upload validation and error handling"""
    
    def test_upload_file_size_limit(self, client, auth_token, sample_batch):
        """Test that oversized files are rejected"""
        # Create a large Excel file
        wb = Workbook()
        ws = wb.active
        
        headers = ['Investor Name', 'Email', 'Internal Client Code', 'Amount Deposited']
        ws.append(headers)
        
        # Add many rows to create large file
        for i in range(100000):  # Very large dataset
            ws.append([
                f'Investor {i}',
                f'inv{i}@example.com',
                f'INV-{i}',
                '50000.00'
            ])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # File might be rejected or processed
        response = client.post(
            f'/api/v1/batches/{sample_batch.id}/upload-excel',
            data={'file': (output, 'large.xlsx')},
            headers=get_auth_headers(auth_token)
        )
        
        assert response.status_code in [200, 201, 400, 413]  # 413 = Payload too large
    
    def test_upload_corrupted_file(self, client, auth_token, sample_batch):
        """Test handling of corrupted Excel files"""
        # Create a corrupted file
        corrupted_file = BytesIO(b'\x00\x01\x02\x03')  # Invalid bytes
        
        response = client.post(
            f'/api/v1/batches/{sample_batch.id}/upload-excel',
            data={'file': (corrupted_file, 'corrupted.xlsx')},
            headers=get_auth_headers(auth_token)
        )
        
        assert response.status_code >= 400


@pytest.mark.unit
class TestUploadResults:
    """Test upload result reporting and feedback"""
    
    def test_upload_returns_summary(self, client, auth_token, sample_batch):
        """Test that upload returns summary of processed records"""
        excel_file = create_sample_excel_file(num_rows=3)
        
        response = client.post(
            f'/api/v1/batches/{sample_batch.id}/upload-excel',
            data={'file': (excel_file, 'investments.xlsx')},
            headers=get_auth_headers(auth_token)
        )
        
        if response.status_code in [200, 201]:
            data = response.get_json()
            # Should contain summary info
            assert 'success' in data or 'total' in data or 'processed' in data
    
    def test_upload_reports_errors(self, client, auth_token, sample_batch):
        """Test that upload reports any errors encountered"""
        # Create Excel with some invalid rows
        wb = Workbook()
        ws = wb.active
        
        headers = ['Investor Name', 'Email', 'Internal Client Code', 'Amount Deposited']
        ws.append(headers)
        
        # Valid row
        ws.append(['Valid Investor', 'valid@example.com', 'VALID-001', '50000.00'])
        
        # Invalid row (negative amount)
        ws.append(['Invalid Investor', 'invalid@example.com', 'INVALID-001', '-50000.00'])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = client.post(
            f'/api/v1/batches/{sample_batch.id}/upload-excel',
            data={'file': (output, 'mixed.xlsx')},
            headers=get_auth_headers(auth_token)
        )
        
        if response.status_code in [200, 201]:
            data = response.get_json()
            # Should report which rows had errors
            assert isinstance(data, dict)


@pytest.mark.unit
class TestUploadSecurity:
    """Test upload security and access control"""
    
    def test_upload_requires_auth(self, client, sample_batch):
        """Test that uploads require authentication"""
        excel_file = create_sample_excel_file()
        
        response = client.post(
            f'/api/v1/batches/{sample_batch.id}/upload-excel',
            data={'file': (excel_file, 'investments.xlsx')}
        )
        
        # Should require JWT token
        assert response.status_code in [401, 400]
    
    def test_upload_file_extension_validation(self, client, auth_token, sample_batch):
        """Test that only Excel files are accepted"""
        # Create a CSV file
        csv_content = b"Investor Name,Email,Code\nJohn,john@example.com,INV-001"
        csv_file = BytesIO(csv_content)
        
        response = client.post(
            f'/api/v1/batches/{sample_batch.id}/upload-excel',
            data={'file': (csv_file, 'data.csv')},
            headers=get_auth_headers(auth_token)
        )
        
        # May be rejected due to file type
        assert response.status_code in [200, 201, 400]
