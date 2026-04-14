import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
from datetime import datetime, timedelta

# Read the investors file
file_path = "investors_test_15_new.xlsx"
df = pd.read_excel(file_path)

print("Source file columns:")
print(df.columns.tolist())
print("\nFirst few rows:")
print(df.head(10))

# Create new workbook for withdrawals
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Withdrawals"

# Define styles
header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
summary_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
summary_font = Font(bold=True, size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
currency_format = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'
percent_format = '0.00%'

# Headers for withdrawal sheet
headers = [
    "Investor Email",
    "Client Code",
    "Deposit Amount (USD)",
    "Valuation (USD)",
    "Profit/Loss (USD)",
    "Profit/Loss (%)",
    "Fund Name",
    "Withdrawal Date",
    "Wealth Manager",
    "IFA",
    "Contract Note"
]

# Add headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Add data rows
row_num = 2
for idx, record in df.iterrows():
    investor_email = record.get('investor_email', '')
    client_code = record.get('internal_client_code', '')
    deposit_amount = float(record.get('amount(usd)', 0))
    valuation = float(record.get('valuation', 0))
    profit_loss = valuation - deposit_amount
    profit_loss_pct = (profit_loss / deposit_amount) if deposit_amount > 0 else 0
    fund_name = record.get('fund_name', '')
    date_deposited = record.get('date_deposited', '')
    wealth_manager = record.get('wealth_manager', '')
    ifa = record.get('IFA', '')
    contract_note = record.get('contract_note', '')
    
    # Withdrawal date is today
    withdrawal_date = datetime.now().strftime('%Y-%m-%d')
    
    row_data = [
        investor_email,
        client_code,
        deposit_amount,
        valuation,
        profit_loss,
        profit_loss_pct,
        fund_name,
        withdrawal_date,
        wealth_manager,
        ifa,
        contract_note
    ]
    
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Format currency columns
        if col_num in [3, 4, 5]:  # Deposit, Valuation, Profit/Loss amounts
            cell.number_format = currency_format
            cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Format percentage column
        if col_num == 6:  # Profit/Loss %
            cell.number_format = percent_format
            cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Format dates
        if col_num == 8:  # Withdrawal Date
            cell.number_format = 'YYYY-MM-DD'
    
    row_num += 1

# Add summary section
summary_row = row_num + 2

# Summary headers
ws.cell(row=summary_row, column=1).value = "WITHDRAWAL SUMMARY"
ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
ws.cell(row=summary_row, column=1).fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

summary_row += 1

# Summary calculations
total_deposit = df['amount(usd)'].sum()
total_valuation = df['valuation'].sum()
total_profit = total_valuation - total_deposit
total_profit_pct = (total_profit / total_deposit) if total_deposit > 0 else 0
investor_count = len(df)
withdrawal_count = len(df)

summary_data = [
    ("Total Investors", investor_count),
    ("Withdrawals Processed", withdrawal_count),
    ("Total Deposited", total_deposit),
    ("Total Valuation", total_valuation),
    ("Total Profit/Loss", total_profit),
    ("Average Return (%)", total_profit_pct),
]

for label, value in summary_data:
    ws.cell(row=summary_row, column=1).value = label
    ws.cell(row=summary_row, column=1).font = summary_font
    ws.cell(row=summary_row, column=1).fill = summary_fill
    ws.cell(row=summary_row, column=1).border = border
    
    cell = ws.cell(row=summary_row, column=2)
    cell.value = value
    cell.font = summary_font
    cell.fill = summary_fill
    cell.border = border
    
    if isinstance(value, float):
        if "%" in label:
            cell.number_format = percent_format
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.number_format = currency_format
            cell.alignment = Alignment(horizontal="right", vertical="center")
    else:
        cell.alignment = Alignment(horizontal="right", vertical="center")
    
    summary_row += 1

# Adjust column widths
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15
ws.column_dimensions['I'].width = 18
ws.column_dimensions['J'].width = 15
ws.column_dimensions['K'].width = 15

# Save the file
output_file = "Withdrawal_Statement.xlsx"
wb.save(output_file)
print(f"\n✅ Withdrawal sheet created: {output_file}")
print(f"Total investors: {investor_count}")
print(f"Total deposited: ${total_deposit:,.2f}")
print(f"Total valuation: ${total_valuation:,.2f}")
print(f"Total profit/loss: ${total_profit:,.2f}")
print(f"Average return: {total_profit_pct:.2%}")
