"""
Generate test Excel file with 15 investors for OFDS testing

This script creates an Excel file with:
- 7 Axiom investors (positions 1-7, auto-assigned)
- 8 Atium investors (positions 8-15, auto-assigned)

Total capital: $600,000
- Axiom: $350,000 (7 × $50,000)
- Atium: $400,000 (8 × $50,000)

Usage:
    python create_test_excel.py

This will create: investors_test_15.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os

# Update this if running from different directory
OUTPUT_FILE = "investors_test_15.xlsx"

def create_test_excel():
    """Create Excel file with 15 test investors"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Investors"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        "investor_name",
        "investor_email",
        "internal_client_code",
        "amount(usd)",
        "fund",
        "date_transferred"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16
    
    # Test data
    investors = [
        # AXIOM FUND (7 investors, positions 1-7)
        ("John Smith", "john.smith@example.com", "AXIOM-001", 50000.00, "Axiom", "2026-03-10"),
        ("Jane Doe", "jane.doe@example.com", "AXIOM-002", 50000.00, "Axiom", "2026-03-10"),
        ("Michael Johnson", "michael.j@example.com", "AXIOM-003", 50000.00, "Axiom", "2026-03-10"),
        ("Sarah Williams", "sarah.w@example.com", "AXIOM-004", 50000.00, "Axiom", "2026-03-10"),
        ("David Brown", "david.b@example.com", "AXIOM-005", 50000.00, "Axiom", "2026-03-10"),
        ("Emma Davis", "emma.d@example.com", "AXIOM-006", 50000.00, "Axiom", "2026-03-10"),
        ("Robert Wilson", "robert.w@example.com", "AXIOM-007", 50000.00, "Axiom", "2026-03-10"),
        
        # ATIUM FUND (8 investors, positions 8-15)
        ("Lisa Anderson", "lisa.a@example.com", "ATIUM-001", 50000.00, "Atium", "2026-03-10"),
        ("James Taylor", "james.t@example.com", "ATIUM-002", 50000.00, "Atium", "2026-03-10"),
        ("Mary Martinez", "mary.m@example.com", "ATIUM-003", 50000.00, "Atium", "2026-03-10"),
        ("William Garcia", "william.g@example.com", "ATIUM-004", 50000.00, "Atium", "2026-03-10"),
        ("Patricia Robinson", "patricia.r@example.com", "ATIUM-005", 50000.00, "Atium", "2026-03-10"),
        ("Christopher Lee", "christopher.l@example.com", "ATIUM-006", 50000.00, "Atium", "2026-03-10"),
        ("Jennifer White", "jennifer.w@example.com", "ATIUM-007", 50000.00, "Atium", "2026-03-10"),
        ("Andrew Harris", "andrew.h@example.com", "ATIUM-008", 50000.00, "Atium", "2026-03-10"),
    ]
    
    # Add data rows
    for row_idx, investor_data in enumerate(investors, 2):
        for col_idx, value in enumerate(investor_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # Format currency column (D)
            if col_idx == 4:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            # Format date column (F)
            elif col_idx == 6:
                cell.number_format = 'YYYY-MM-DD'
                cell.alignment = center_align
            # Center align other columns
            else:
                cell.alignment = Alignment(horizontal="left")
    
    # Add summary section
    summary_row = len(investors) + 3
    
    # Fund summary
    ws.cell(row=summary_row, column=1, value="SUMMARY")
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=11)
    
    ws.cell(row=summary_row + 1, column=1, value="Fund")
    ws.cell(row=summary_row + 1, column=2, value="Investor Count")
    ws.cell(row=summary_row + 1, column=3, value="Total Capital")
    
    for cell in ws[summary_row + 1]:
        if cell.value:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    ws.cell(row=summary_row + 2, column=1, value="Axiom")
    ws.cell(row=summary_row + 2, column=2, value=7)
    ws.cell(row=summary_row + 2, column=3, value=350000.00)
    ws.cell(row=summary_row + 2, column=3).number_format = '$#,##0.00'
    
    ws.cell(row=summary_row + 3, column=1, value="Atium")
    ws.cell(row=summary_row + 3, column=2, value=8)
    ws.cell(row=summary_row + 3, column=3, value=400000.00)
    ws.cell(row=summary_row + 3, column=3).number_format = '$#,##0.00'
    
    ws.cell(row=summary_row + 4, column=1, value="TOTAL")
    ws.cell(row=summary_row + 4, column=1).font = Font(bold=True)
    ws.cell(row=summary_row + 4, column=2, value=15)
    ws.cell(row=summary_row + 4, column=2).font = Font(bold=True)
    ws.cell(row=summary_row + 4, column=3, value=750000.00)
    ws.cell(row=summary_row + 4, column=3).number_format = '$#,##0.00'
    ws.cell(row=summary_row + 4, column=3).font = Font(bold=True)
    
    # Save file
    wb.save(OUTPUT_FILE)
    print(f"✅ Excel file created: {OUTPUT_FILE}")
    print(f"   Location: {os.path.abspath(OUTPUT_FILE)}")
    print(f"\n📊 Contents:")
    print(f"   - Axiom: 7 investors × $50,000 = $350,000")
    print(f"   - Atium: 8 investors × $50,000 = $400,000")
    print(f"   - TOTAL: 15 investors × $50,000 = $750,000")
    print(f"\n✨ Ready for testing with your OFDS API!")

if __name__ == "__main__":
    create_test_excel()
