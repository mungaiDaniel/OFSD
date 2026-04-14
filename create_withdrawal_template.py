import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
import pandas as pd
from datetime import datetime

# Read the investors file
file_path = "investors_test_15_new.xlsx"
df = pd.read_excel(file_path)

# Filter out summary rows (non-numeric data)
df = df[df['investor_name'].notna() & (df['investor_name'] != 'SUMMARY') & (df['investor_name'] != 'Fund')]
df = df.reset_index(drop=True)

# Create new workbook for withdrawals with CORRECT Withdrawal model columns
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Withdrawals"

# Define styles
header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
required_header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
required_font = Font(bold=True, color="FFFFFF", size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
currency_format = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'
date_format = 'YYYY-MM-DD'

# Headers matching Withdrawal model columns
# Required columns (marked RED)
# Optional columns (marked BLUE)
headers = [
    ("internal_client_code", True),      # Required - from Withdrawal model
    ("investor_name", True),              # Required - reference only
    ("fund_name", True),                  # Required - from Withdrawal model
    ("amount", True),                     # Required - from Withdrawal model
    ("date_withdrawn", True),             # Required - from Withdrawal model
    ("status", False),                    # Optional - defaults to 'Pending'
    ("note", False),                      # Optional - any withdrawal notes
]

# Add headers with color coding
for col_num, (header, is_required) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = required_font if is_required else header_font
    cell.fill = required_header_fill if is_required else header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    
    # Add comments to indicate required/optional
    if is_required:
        cell.comment = Comment("Required field", "System")
    else:
        cell.comment = Comment("Optional - system will use default if blank", "System")

# Add data rows (starting from row 2)
row_num = 2
for idx, record in df.iterrows():
    investor_name = str(record.get('investor_name', '')).strip()
    if not investor_name or investor_name.upper() in ['NAN', '']:
        continue
    
    internal_code = record.get('internal_client_code', '')
    fund_name = record.get('fund_name', '')
    amount = float(record.get('amount(usd)', 0))
    
    # date_withdrawn from investment's date_deposited (when they invested)
    withdrawal_date = datetime.now().strftime('%Y-%m-%d')
    
    # Status defaults to 'Pending' but can be set to 'Approved' or 'Rejected'
    status = 'Pending'
    
    # Note field - optional
    note = ''
    
    row_data = [
        internal_code,
        investor_name,
        fund_name,
        amount,
        withdrawal_date,
        status,
        note
    ]
    
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Format currency column (amount)
        if col_num == 4:  # amount
            cell.number_format = currency_format
            cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Format date column (date_withdrawn)
        if col_num == 5:  # date_withdrawn
            cell.number_format = date_format
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Center align status
        if col_num == 6:  # status
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row_num += 1

# Column widths
ws.column_dimensions['A'].width = 20  # internal_client_code
ws.column_dimensions['B'].width = 25  # investor_name
ws.column_dimensions['C'].width = 15  # fund_name
ws.column_dimensions['D'].width = 18  # amount
ws.column_dimensions['E'].width = 18  # date_withdrawn
ws.column_dimensions['F'].width = 15  # status
ws.column_dimensions['G'].width = 30  # note

# Freeze header row only
ws.freeze_panes = ws['A2']

# Save the file
output_file = "Withdrawal_Statement.xlsx"
wb.save(output_file)

# Print summary
total_withdrawn = df['amount(usd)'].sum()
withdrawal_count = len(df)

print(f"✅ Withdrawal template created: {output_file}")
print(f"\nWithdrawal Model Database Schema:")
print(f"  ✓ id (auto-generated)")
print(f"  ✓ internal_client_code (REQUIRED)")
print(f"  ✓ fund_id (optional, resolved from fund_name)")
print(f"  ✓ fund_name (REQUIRED)")
print(f"  ✓ amount (REQUIRED)")
print(f"  ✓ date_withdrawn (REQUIRED)")
print(f"  ✓ status (optional, defaults to 'Pending')")
print(f"  ✓ approved_at (optional)")
print(f"  ✓ note (optional)")
print(f"  ✓ batch_id (auto-linked from investment)")
print(f"\nExcel Columns (RED = Required, BLUE = Optional):")
print(f"  [RED]  internal_client_code")
print(f"  [RED]  investor_name")
print(f"  [RED]  fund_name")
print(f"  [RED]  amount")
print(f"  [RED]  date_withdrawn")
print(f"  [BLUE] status")
print(f"  [BLUE] note")
print(f"\nWithdrawal Summary:")
print(f"  Total investors: {withdrawal_count}")
print(f"  Total amount: ${total_withdrawn:,.2f}")
print(f"  Withdrawal date: {datetime.now().strftime('%Y-%m-%d')}")
print(f"\n📋 All data aligned with Withdrawal model columns")
print(f"✅ Ready to upload via /withdrawals/upload endpoint")
